"""
GrabFood Menu Scraper — Streamlit App (IMPROVED v2)
====================================================
Perbaikan utama:
  - Multi-strategy selector dengan fallback berlapis
  - Anti-bot detection yang lebih baik (stealth mode)
  - Nama restoran dari berbagai sumber
  - Scraping berbasis teks + regex sebagai fallback terakhir
  - Logging debug opsional untuk troubleshooting
  - Retry otomatis jika halaman gagal load
"""

import io
import time
import re
import json

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


# ─────────────────────────────────────────────
# AUTO-INSTALL PLAYWRIGHT BROWSER
# ─────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def install_playwright_browsers():
    import subprocess, sys
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        check=True, capture_output=True,
    )

install_playwright_browsers()


# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="GrabFood Menu Scraper",
    page_icon="🍔",
    layout="centered",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main-title { text-align:center; font-size:2.2rem; font-weight:700; color:#00b14f; margin-bottom:0.2rem; }
    .sub-title { text-align:center; font-size:1rem; color:#555; margin-bottom:2rem; }
    .info-box { background:#f0fdf4; border-left:4px solid #00b14f; border-radius:6px; padding:0.8rem 1.2rem; margin-bottom:1.5rem; font-size:0.9rem; color:#166534; }
    .resto-header { background:#f8f9fa; border-radius:8px; padding:0.6rem 1rem; margin:1rem 0 0.4rem 0; font-weight:600; font-size:1rem; color:#111; border-left:4px solid #00b14f; }
    .stat-card { background:#f8f9fa; border-radius:10px; padding:1rem 1.2rem; text-align:center; border:1px solid #e9ecef; }
    .stat-card .label { font-size:0.75rem; color:#888; font-weight:500; text-transform:uppercase; letter-spacing:0.05em; }
    .stat-card .value { font-size:1.3rem; font-weight:700; color:#00b14f; margin-top:0.2rem; }
    div[data-testid="stButton"] > button { background:#00b14f; color:white; border:none; border-radius:8px; padding:0.6rem 2rem; font-size:1rem; font-weight:600; width:100%; }
    div[data-testid="stButton"] > button:hover { background:#009942; }
    div[data-testid="stDownloadButton"] > button { background:#1d4ed8; color:white; border:none; border-radius:8px; padding:0.6rem 2rem; font-size:1rem; font-weight:600; width:100%; }
    div[data-testid="stDownloadButton"] > button:hover { background:#1e40af; }
    .debug-box { background:#1e1e1e; color:#a8ff78; font-family:monospace; font-size:0.78rem; border-radius:6px; padding:0.8rem 1rem; max-height:200px; overflow-y:auto; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# BROWSER CONTEXT FACTORY (stealth)
# ─────────────────────────────────────────────
def create_stealth_context(browser):
    """Buat context browser dengan setting anti-deteksi."""
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1366, "height": 768},
        locale="id-ID",
        timezone_id="Asia/Jakarta",
        extra_http_headers={
            "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        },
        java_script_enabled=True,
    )
    # Sembunyikan properti webdriver
    context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        Object.defineProperty(navigator, 'languages', { get: () => ['id-ID', 'id', 'en-US', 'en'] });
        window.chrome = { runtime: {} };
    """)
    return context


# ─────────────────────────────────────────────
# AMBIL NAMA RESTORAN (multi-strategy)
# ─────────────────────────────────────────────
def get_restaurant_name(page, url):
    """Coba berbagai cara untuk mendapatkan nama restoran."""
    
    # Strategi 1: Selector spesifik GrabFood
    selectors = [
        '[data-testid="restaurant-name"]',
        '[class*="restaurantName"]',
        '[class*="restaurant-name"]',
        '[class*="headerName"]',
        '[class*="RestaurantHeader"] h1',
        '[class*="RestaurantHeader"] h2',
        '.restaurant-name',
        '#restaurant-name',
        'h1[class*="name"]',
        'h1[class*="Name"]',
        'h1',
        'h2',
    ]
    for sel in selectors:
        try:
            elem = page.query_selector(sel)
            if elem:
                name = elem.inner_text().strip()
                # Filter nama yang masuk akal (bukan UI label)
                if name and len(name) > 2 and len(name) < 100:
                    if not any(skip in name.lower() for skip in ['grabfood', 'home', 'back', 'menu', 'cart']):
                        return name
        except Exception:
            continue

    # Strategi 2: Dari title tag
    try:
        title = page.title()
        if title:
            for sep in [' | ', ' - ', ' – ', ' — ']:
                if sep in title:
                    name = title.split(sep)[0].strip()
                    if name and 'grab' not in name.lower():
                        return name
    except Exception:
        pass

    # Strategi 3: Dari URL slug
    try:
        # Ambil slug restoran dari URL: /restaurant/SLUG-delivery/ID
        match = re.search(r'/restaurant/([^/]+?)(?:-delivery)?/[^/]+$', url)
        if match:
            slug = match.group(1)
            # Ubah slug jadi nama: "mie-aceh-bang-ibra" → "Mie Aceh Bang Ibra"
            name = ' '.join(word.capitalize() for word in slug.split('-'))
            return name
    except Exception:
        pass

    # Strategi 4: Dari JSON-LD / structured data
    try:
        scripts = page.query_selector_all('script[type="application/ld+json"]')
        for script in scripts:
            try:
                data = json.loads(script.inner_text())
                if isinstance(data, dict) and data.get('name'):
                    return data['name']
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict) and item.get('name'):
                            return item['name']
            except Exception:
                continue
    except Exception:
        pass

    return "Restoran"


# ─────────────────────────────────────────────
# SCROLL HALAMAN
# ─────────────────────────────────────────────
def scroll_to_load(page, max_scrolls=15, wait_ms=3000, progress_cb=None):
    """Scroll halaman sambil menunggu konten dinamis."""
    last_count = 0
    no_change = 0

    for i in range(max_scrolls):
        # Scroll bertahap
        page.evaluate("window.scrollBy(0, window.innerHeight * 0.8)")
        page.wait_for_timeout(wait_ms)

        # Hitung item menu yang sudah ada
        current_count = page.evaluate("""
            () => document.querySelectorAll('[class*="menuItem"], [class*="MenuItem"], [class*="menu-item"]').length
        """)

        if current_count == last_count:
            no_change += 1
            if no_change >= 3:
                break
        else:
            no_change = 0
            last_count = current_count

        if progress_cb:
            progress_cb(i + 1, max_scrolls)

    # Scroll balik ke atas
    page.evaluate("window.scrollTo(0, 0)")
    page.wait_for_timeout(1500)


# ─────────────────────────────────────────────
# EXTRACT HARGA
# ─────────────────────────────────────────────
def extract_price(text):
    if not text:
        return None
    digits = re.sub(r'[^\d]', '', str(text))
    if digits and len(digits) >= 3:  # min Rp 100
        val = int(digits)
        if 100 <= val <= 10_000_000:  # sanity check
            return val
    return None


def find_price_in_text(text):
    """Cari pola harga IDR dalam teks bebas."""
    patterns = [
        r'Rp\s*([\d.,]+)',
        r'IDR\s*([\d.,]+)',
        r'([\d]{2,3}(?:[.,]\d{3})+)',  # 25.000 atau 25,000
        r'([\d]{4,6})',                  # 5000-999999
    ]
    for pat in patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        for m in matches:
            price = extract_price(m)
            if price and price >= 1000:
                return price
    return None


# ─────────────────────────────────────────────
# SCRAPE MENU — STRATEGY 1: DOM Selector
# ─────────────────────────────────────────────
def scrape_via_selectors(page):
    """Coba berbagai selector CSS untuk item menu."""
    
    # Kumpulan selector container menu item
    container_selectors = [
        '[class*="menuItem"]',
        '[class*="MenuItem"]',
        '[class*="menu-item"]',
        '[class*="FoodItem"]',
        '[class*="foodItem"]',
        '[class*="product-item"]',
        '[class*="ProductItem"]',
        '[class*="itemCard"]',
        '[class*="ItemCard"]',
        'div.ant-row[class*="item"]',
        'div.ant-row',
    ]

    items = []
    for sel in container_selectors:
        try:
            elements = page.query_selector_all(sel)
            if len(elements) >= 3:  # minimal 3 item → selector ini valid
                items = elements
                break
        except Exception:
            continue

    if not items:
        return []

    hasil = []
    seen_names = set()

    for item in items:
        try:
            full_text = item.inner_text().strip()
            if not full_text or len(full_text) < 3:
                continue

            # ── Nama menu ──
            name = ''
            name_selectors = [
                '[class*="itemName"]', '[class*="ItemName"]',
                '[class*="menuName"]', '[class*="MenuName"]',
                '[class*="productName"]', '[class*="ProductName"]',
                '[class*="name"]', '[class*="title"]', '[class*="Title"]',
                'h3', 'h4', 'p[class*="name"]',
            ]
            for ns in name_selectors:
                try:
                    elem = item.query_selector(ns)
                    if elem:
                        t = elem.inner_text().strip()
                        if t and len(t) > 1 and not re.search(r'^Rp|^\d', t):
                            name = t
                            break
                except Exception:
                    continue

            # Fallback: baris pertama teks
            if not name:
                lines = [l.strip() for l in full_text.split('\n') if l.strip()]
                for line in lines:
                    if line and not re.search(r'^Rp|^\d{3,}', line):
                        name = line
                        break

            if not name or len(name) < 2:
                continue

            # De-duplikasi
            name_key = name.lower().strip()
            if name_key in seen_names:
                continue
            seen_names.add(name_key)

            # ── Deskripsi ──
            desc = ''
            desc_selectors = [
                '[class*="description"]', '[class*="Description"]',
                '[class*="desc"]', '[class*="subtitle"]',
            ]
            for ds in desc_selectors:
                try:
                    elem = item.query_selector(ds)
                    if elem:
                        t = elem.inner_text().strip()
                        if t and t != name:
                            desc = t[:200]
                            break
                except Exception:
                    continue

            # Fallback deskripsi dari teks
            if not desc:
                lines = [l.strip() for l in full_text.split('\n') if l.strip()]
                non_price = [l for l in lines[1:] if not re.search(r'^Rp|^\d{4,}', l)]
                if non_price:
                    desc = ' '.join(non_price[:2])[:200]

            # ── Harga ──
            price = None
            price_selectors = [
                '[class*="discountedPrice"]', '[class*="DiscountedPrice"]',
                '[class*="finalPrice"]', '[class*="FinalPrice"]',
                '[class*="itemPrice"]', '[class*="ItemPrice"]',
                '[class*="price"]', '[class*="Price"]',
            ]
            for ps in price_selectors:
                try:
                    elem = item.query_selector(ps)
                    if elem:
                        price = extract_price(elem.inner_text())
                        if price:
                            break
                except Exception:
                    continue

            # Fallback harga dari full text
            if not price:
                price = find_price_in_text(full_text)

            hasil.append({
                'Nama Menu': name,
                'Deskripsi': desc,
                'Harga (Rp)': price,
                'Harga': f"Rp {price:,}" if price else 'N/A',
            })

        except Exception:
            continue

    return hasil


# ─────────────────────────────────────────────
# SCRAPE MENU — STRATEGY 2: JSON dari window/__NEXT_DATA__
# ─────────────────────────────────────────────
def scrape_via_json(page):
    """Coba ekstrak data dari __NEXT_DATA__ atau window.__DATA__ GrabFood."""
    hasil = []
    try:
        # Coba __NEXT_DATA__
        next_data = page.evaluate("""
            () => {
                const el = document.getElementById('__NEXT_DATA__');
                return el ? el.textContent : null;
            }
        """)
        if next_data:
            data = json.loads(next_data)
            # Traversal rekursif cari array item menu
            items = find_menu_items_in_json(data)
            if items:
                return items
    except Exception:
        pass

    try:
        # Coba ambil dari network cache via performance entries
        resources = page.evaluate("""
            () => {
                const entries = performance.getEntriesByType('resource');
                return entries
                    .filter(e => e.name.includes('api') || e.name.includes('menu'))
                    .map(e => e.name)
                    .slice(0, 10);
            }
        """)
    except Exception:
        pass

    return hasil


def find_menu_items_in_json(obj, depth=0):
    """Rekursif cari struktur menu dalam JSON."""
    if depth > 10:
        return []
    
    hasil = []
    
    if isinstance(obj, list):
        for item in obj:
            if isinstance(item, dict):
                # Deteksi item menu
                name = item.get('name') or item.get('itemName') or item.get('title', '')
                price_raw = (item.get('price') or item.get('displayedPrice') or 
                             item.get('priceInMin') or item.get('priceDisplay') or '')
                
                if name and len(str(name)) > 1:
                    price = extract_price(str(price_raw)) if price_raw else None
                    desc = str(item.get('description', '') or '')[:200]
                    hasil.append({
                        'Nama Menu': str(name),
                        'Deskripsi': desc,
                        'Harga (Rp)': price,
                        'Harga': f"Rp {price:,}" if price else 'N/A',
                    })
                else:
                    hasil.extend(find_menu_items_in_json(item, depth + 1))
            elif isinstance(item, (list, dict)):
                hasil.extend(find_menu_items_in_json(item, depth + 1))
    
    elif isinstance(obj, dict):
        for key, val in obj.items():
            if key in ('items', 'menuItems', 'products', 'dishes', 'foods', 'menu'):
                if isinstance(val, list) and len(val) > 0:
                    found = find_menu_items_in_json(val, depth + 1)
                    if found:
                        return found
            elif isinstance(val, (dict, list)):
                found = find_menu_items_in_json(val, depth + 1)
                if found:
                    hasil.extend(found)
    
    return hasil


# ─────────────────────────────────────────────
# SCRAPE MENU — STRATEGY 3: Full page text parsing
# ─────────────────────────────────────────────
def scrape_via_text_parsing(page):
    """
    Last resort: parsing teks mentah halaman.
    Cari pola: [nama menu] [deskripsi opsional] [Rp xxx]
    """
    try:
        body_text = page.inner_text('body')
    except Exception:
        return []

    lines = [l.strip() for l in body_text.split('\n') if l.strip()]
    hasil = []
    seen = set()
    i = 0

    while i < len(lines):
        line = lines[i]
        
        # Cek apakah baris berikutnya atau baris ini mengandung harga
        price = None
        name = ''
        desc = ''

        # Pola: baris ini = nama, baris berikut = harga
        if i + 1 < len(lines):
            next_line = lines[i + 1]
            price = find_price_in_text(next_line)
            if price and not re.search(r'^Rp|^\d', line) and len(line) > 2:
                name = line
                # Cek deskripsi (baris setelah nama, sebelum harga)
                if not price:  # tidak langsung harga
                    desc = next_line[:150]
                i += 2
            else:
                # Cek apakah baris ini langsung ada harga
                inline_price = find_price_in_text(line)
                if inline_price:
                    # Baris ini gabungan nama + harga, skip
                    i += 1
                    continue
                i += 1
                continue
        else:
            i += 1
            continue

        if not name or len(name) < 3:
            continue
        
        # Filter baris yang bukan nama menu
        skip_patterns = [
            r'^(home|back|cart|search|order|delivery|pickup|promo|voucher|lihat|tambah|pilih)',
            r'^\d+$',
            r'^(rp|idr)',
            r'(restoran|restaurant|grab|food)',
        ]
        if any(re.search(p, name, re.IGNORECASE) for p in skip_patterns):
            continue

        name_key = name.lower()
        if name_key in seen:
            continue
        seen.add(name_key)

        hasil.append({
            'Nama Menu': name,
            'Deskripsi': desc,
            'Harga (Rp)': price,
            'Harga': f"Rp {price:,}" if price else 'N/A',
        })

    return hasil


# ─────────────────────────────────────────────
# MAIN SCRAPER
# ─────────────────────────────────────────────
def scrape_menu(page, url, progress_cb=None, debug_log=None):
    """
    Scrape menu dengan multi-strategy + retry.
    Returns (resto_name, menu_list, strategy_used)
    """
    def log(msg):
        if debug_log is not None:
            debug_log.append(msg)

    # ── Load halaman dengan retry ──────────────────────────────────────────
    for attempt in range(2):
        try:
            log(f"[{attempt+1}] Loading: {url}")
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(3000)
            
            # Tunggu sampai konten muncul (maksimal 15 detik)
            try:
                page.wait_for_selector(
                    '[class*="menu"], [class*="Menu"], [class*="item"], [class*="Item"]',
                    timeout=15000
                )
            except Exception:
                log("  Timeout menunggu selector menu, lanjut scroll...")
            break
        except Exception as e:
            log(f"  Error load: {e}")
            if attempt == 1:
                return "Restoran", [], "error"
            time.sleep(3)

    # ── Ambil nama restoran ────────────────────────────────────────────────
    resto_name = get_restaurant_name(page, url)
    log(f"  Nama restoran: {resto_name}")

    # ── Scroll untuk load konten ───────────────────────────────────────────
    scroll_to_load(page, max_scrolls=15, wait_ms=3000, progress_cb=progress_cb)

    # ── Strategy 1: DOM Selectors ──────────────────────────────────────────
    log("  Strategy 1: DOM selectors...")
    menu_list = scrape_via_selectors(page)
    log(f"  → {len(menu_list)} item")
    
    if len(menu_list) >= 3:
        return resto_name, menu_list, "DOM selectors"

    # ── Strategy 2: JSON / __NEXT_DATA__ ──────────────────────────────────
    log("  Strategy 2: JSON data...")
    json_list = scrape_via_json(page)
    log(f"  → {len(json_list)} item")
    
    if len(json_list) > len(menu_list):
        menu_list = json_list
    
    if len(menu_list) >= 3:
        return resto_name, menu_list, "JSON data"

    # ── Strategy 3: Text parsing ───────────────────────────────────────────
    log("  Strategy 3: Text parsing...")
    text_list = scrape_via_text_parsing(page)
    log(f"  → {len(text_list)} item")
    
    if len(text_list) > len(menu_list):
        menu_list = text_list
        strategy = "text parsing"
    else:
        strategy = "DOM selectors (partial)"

    return resto_name, menu_list, strategy


# ─────────────────────────────────────────────
# DE-DUPLIKASI
# ─────────────────────────────────────────────
def remove_duplicates(menu_list):
    seen = set()
    unique = []
    for item in menu_list:
        key = re.sub(r'\s+', ' ', item['Nama Menu'].strip().lower())
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
BLACK_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN_FILL = PatternFill("solid", fgColor="00b14f")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")


def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = BLACK_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_header(ws, header_row=1, fill=GREEN_FILL, font_color="FFFFFF"):
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = fill


def safe_sheet_name(name, existing_names):
    name = re.sub(r'[\\/*?:\[\]]', '', name)[:31]
    if not name:
        name = "Sheet"
    base = name
    counter = 1
    while name in existing_names:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    return name


def build_excel(results):
    output = io.BytesIO()
    summary_rows = []
    sheet_names = []

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for resto_name, url, menu_list, strategy in results:
            _cols = ['Nama Menu', 'Deskripsi', 'Harga']
            df = pd.DataFrame(menu_list, columns=_cols) if menu_list else pd.DataFrame(columns=_cols)
            sheet_name = safe_sheet_name(resto_name, sheet_names)
            sheet_names.append(sheet_name)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

            prices = [m['Harga (Rp)'] for m in menu_list if m.get('Harga (Rp)')]
            summary_rows.append({
                'Restoran': resto_name,
                'URL': url,
                'Total Menu': len(menu_list),
                'Strategi': strategy,
                'Rata-rata Harga': f"Rp {sum(prices)//len(prices):,}" if prices else 'N/A',
                'Harga Terendah': f"Rp {min(prices):,}" if prices else 'N/A',
                'Harga Tertinggi': f"Rp {max(prices):,}" if prices else 'N/A',
            })

        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="SUMMARY", index=False)

    output.seek(0)
    wb = load_workbook(output)

    for i, (resto_name, url, menu_list, _) in enumerate(results):
        ws = wb[sheet_names[i]]
        ws["A1"] = "URL"
        ws["B1"] = url
        ws["A1"].font = Font(bold=True)
        apply_borders(ws)
        style_header(ws, header_row=3)
        auto_fit_columns(ws)
        ws.freeze_panes = "A4"

    ws_sum = wb["SUMMARY"]
    apply_borders(ws_sum)
    style_header(ws_sum, header_row=1, fill=GRAY_FILL, font_color="000000")
    auto_fit_columns(ws_sum)
    ws_sum.freeze_panes = "A2"
    wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames) + 1)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


# ─────────────────────────────────────────────
# STAT CARD
# ─────────────────────────────────────────────
def stat_card(label, value, small=False):
    font_size = "0.95rem" if small else "1.3rem"
    return f"""
    <div class="stat-card">
        <div class="label">{label}</div>
        <div class="value" style="font-size:{font_size}">{value}</div>
    </div>"""


# ─────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────
def main():
    st.markdown('<div class="main-title">GrabFood Menu Scraper</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-title">Masukkan satu atau beberapa link restoran GrabFood '
        'dan dapatkan daftar menu dalam format Excel</div>',
        unsafe_allow_html=True
    )

    st.markdown("""
    <div class="info-box">
        📌 <strong>Cara pakai:</strong> Tempel satu atau beberapa URL GrabFood (satu URL per baris),
        lalu klik <strong>Mulai Scraping</strong>.<br>
        ⏱️ Setiap restoran membutuhkan sekitar 1–3 menit.<br>
        🔄 Scraper menggunakan 3 strategi otomatis jika satu tidak berhasil.
    </div>
    """, unsafe_allow_html=True)

    urls_input = st.text_area(
        "🔗 Link Restoran GrabFood (satu URL per baris)",
        placeholder="https://food.grab.com/id/id/restaurant/...\nhttps://food.grab.com/id/id/restaurant/...",
        height=130,
    )

    col_btn, col_debug = st.columns([2, 1])
    with col_btn:
        run_btn = st.button("🚀 Mulai Scraping")
    with col_debug:
        show_debug = st.toggle("🐛 Debug log", value=False)

    if not run_btn:
        return

    raw_urls = [u.strip() for u in urls_input.strip().splitlines() if u.strip()]
    if not raw_urls:
        st.error("⚠️ Masukkan minimal satu URL.")
        return

    invalid = [u for u in raw_urls if "grab.com" not in u]
    if invalid:
        st.warning(f"⚠️ URL berikut mungkin bukan URL GrabFood:\n" + "\n".join(f"- {u}" for u in invalid))

    overall_text = st.empty()
    overall_bar = st.progress(0)
    detail_text = st.empty()
    debug_container = st.empty()

    results = []
    failed_urls = []
    total = len(raw_urls)
    all_debug_logs = []

    with sync_playwright() as p:
        for idx, url in enumerate(raw_urls):
            overall_text.info(f"🔄 Scraping restoran **{idx + 1} / {total}**...")
            overall_bar.progress(int(idx / total * 100))

            debug_log = []

            def scroll_progress(cur, mx, _idx=idx, _total=total):
                inner_pct = int((cur / mx) * 70)
                outer_pct = int((_idx / _total * 100) + (inner_pct / _total))
                overall_bar.progress(min(outer_pct, 99))
                detail_text.caption(f"  ↳ Scroll {cur}/{mx} — memuat konten menu...")

            detail_text.caption("  ↳ Membuka browser...")
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-extensions",
                    "--no-first-run",
                    "--disable-default-apps",
                ],
            )
            try:
                context = create_stealth_context(browser)
                page = context.new_page()

                detail_text.caption("  ↳ Membuka halaman GrabFood...")
                resto_name, menu_list, strategy = scrape_menu(
                    page, url, progress_cb=scroll_progress, debug_log=debug_log
                )
                menu_list = remove_duplicates(menu_list)
                results.append((resto_name, url, menu_list, strategy))

                status_emoji = "✅" if menu_list else "⚠️"
                detail_text.caption(
                    f"  {status_emoji} {resto_name} — {len(menu_list)} item "
                    f"(via {strategy})"
                )

                all_debug_logs.extend([f"[URL {idx+1}] " + l for l in debug_log])

                if show_debug and debug_log:
                    debug_container.markdown(
                        '<div class="debug-box">' +
                        '<br>'.join(debug_log) +
                        '</div>',
                        unsafe_allow_html=True
                    )

            except Exception as e:
                failed_urls.append(url)
                detail_text.caption(f"  ❌ Gagal: {e}")
                all_debug_logs.append(f"[URL {idx+1}] ERROR: {e}")
            finally:
                browser.close()

    overall_bar.progress(100)
    overall_text.success(f"✅ Selesai! {len(results)}/{total} restoran berhasil di-scrape.")

    if show_debug and all_debug_logs:
        with st.expander("🐛 Full Debug Log"):
            st.code('\n'.join(all_debug_logs), language=None)

    if failed_urls:
        st.warning("⚠️ URL berikut gagal:\n" + "\n".join(f"- {u}" for u in failed_urls))

    if not results:
        st.error("❌ Tidak ada data yang berhasil di-scrape.")
        return

    # ── Tampilkan hasil ────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Hasil Scraping")

    total_items_all = sum(len(m) for _, _, m, _ in results)
    
    if len(results) > 1:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(stat_card("Total Restoran", len(results)), unsafe_allow_html=True)
        with c2:
            st.markdown(stat_card("Total Item Menu", total_items_all), unsafe_allow_html=True)
        st.markdown("")

    for resto_name, url, menu_list, strategy in results:
        st.markdown(f'<div class="resto-header">🏪 {resto_name}</div>', unsafe_allow_html=True)

        if not menu_list:
            st.warning(
                f"⚠️ Tidak ada item menu yang berhasil di-scrape untuk **{resto_name}**. "
                f"GrabFood mungkin memblok permintaan ini atau struktur halaman berubah. "
                f"Coba buka URL di browser dan pastikan menu terlihat."
            )
            continue

        prices = [m['Harga (Rp)'] for m in menu_list if m.get('Harga (Rp)')]
        avg = f"Rp {sum(prices)//len(prices):,}" if prices else "N/A"
        low = f"Rp {min(prices):,}" if prices else "N/A"
        high = f"Rp {max(prices):,}" if prices else "N/A"

        st.caption(f"📡 Strategi: {strategy}")

        c1, c2, c3, c4 = st.columns(4)
        with c1: st.markdown(stat_card("Total Menu", len(menu_list)), unsafe_allow_html=True)
        with c2: st.markdown(stat_card("Rata-rata", avg, small=True), unsafe_allow_html=True)
        with c3: st.markdown(stat_card("Terendah", low, small=True), unsafe_allow_html=True)
        with c4: st.markdown(stat_card("Tertinggi", high, small=True), unsafe_allow_html=True)

        with st.expander(f"📋 Preview menu {resto_name} ({len(menu_list)} item)"):
            _cols = ['Nama Menu', 'Deskripsi', 'Harga']
            df = pd.DataFrame(menu_list, columns=_cols)
            st.dataframe(df, use_container_width=True, height=300)

    # ── Download Excel ─────────────────────────────────────────────────────
    st.markdown("---")
    valid_results = [(r, u, m, s) for r, u, m, s in results if m]
    if valid_results:
        with st.spinner("📊 Membuat file Excel..."):
            excel_bytes = build_excel(valid_results)

        n = len(valid_results)
        filename = "grabfood_menu.xlsx" if n == 1 else f"grabfood_menu_{n}_resto.xlsx"

        st.download_button(
            label=f"⬇️ Download Excel ({n} restoran, {total_items_all} item)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.error("❌ Tidak ada data yang cukup untuk dibuat Excel.")


if __name__ == "__main__":
    main()