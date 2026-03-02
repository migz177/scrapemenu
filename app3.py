"""
GrabFood Menu Scraper — Versi Playwright (VPS)
===============================================
Menggunakan Playwright untuk membuka browser Chromium sungguhan di background,
intercept network request ke portal.grab.com, dan ambil data menu secara otomatis.

Setup VPS:
    pip install streamlit playwright pandas openpyxl
    playwright install chromium
    playwright install-deps chromium
"""

import io
import re
import time
import asyncio
import urllib.request
import urllib.parse
import json as _json

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

# ── Install Chromium binary otomatis (wajib untuk Streamlit Cloud) ──────────
import subprocess, sys

@st.cache_resource(show_spinner=False)
def _install_playwright_browsers():
    """Jalankan 'playwright install chromium' sekali saat startup."""
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True, capture_output=True,
        )
    except Exception as e:
        st.warning(f"⚠️ Gagal install Chromium: {e}")

_install_playwright_browsers()

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="GrabFood Menu Scraper",
    page_icon="🍔",
    layout="centered",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main-title  { font-size:2rem; font-weight:700; color:#00b14f; margin-bottom:0.2rem; }
    .sub-title   { font-size:1rem; color:#555; margin-bottom:1.5rem; }
    .info-box    { background:#f0faf4; border-left:4px solid #00b14f; border-radius:6px; padding:.75rem 1rem; margin-bottom:1rem; font-size:.9rem; color:#1a1a1a; }
    .warn-box    { background:#fff8e1; border-left:4px solid #f59e0b; border-radius:6px; padding:.75rem 1rem; margin-bottom:1rem; font-size:.9rem; color:#1a1a1a; }
    .stat-card   { background:#f8f9fa; border:1px solid #e0e0e0; border-radius:10px; padding:.8rem 1rem; text-align:center; margin:.2rem; }
    .stat-card .label { font-size:.75rem; color:#666; font-weight:500; text-transform:uppercase; letter-spacing:.04em; }
    .stat-card .value { font-size:1.3rem; font-weight:700; color:#00b14f; margin-top:.2rem; }
    .resto-header { font-size:1.2rem; font-weight:700; color:#1a1a1a; margin:1.2rem 0 .5rem; border-bottom:2px solid #00b14f; padding-bottom:.3rem; }
    div[data-testid="stButton"]         > button { background:#00b14f; color:white; border:none; border-radius:8px; padding:.6rem 2rem; font-size:1rem; font-weight:600; width:100%; }
    div[data-testid="stButton"]         > button:hover { background:#009942; }
    div[data-testid="stDownloadButton"] > button { background:#1d4ed8; color:white; border:none; border-radius:8px; padding:.6rem 2rem; font-size:1rem; font-weight:600; width:100%; }
    div[data-testid="stDownloadButton"] > button:hover { background:#1e40af; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# PLAYWRIGHT SCRAPER
# ─────────────────────────────────────────────
async def scrape_with_playwright(
    url: str,
    lat: float = -6.2088,
    lng: float = 106.8456,
    timeout_ms: int = 30_000,
) -> dict:
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--no-first-run",
                "--no-zygote",
                "--single-process",
                "--disable-extensions",
            ],
        )
        context = await browser.new_context(
            viewport={"width": 1280, "height": 800},
            locale="id-ID",
            timezone_id="Asia/Jakarta",
            geolocation={"latitude": lat, "longitude": lng},
            permissions=["geolocation"],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
            extra_http_headers={"Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8"},
        )
        page = await context.new_page()
        captured_data = {}

        async def handle_response(response):
            resp_url = response.url
            if (
                "portal.grab.com" in resp_url
                and "merchants" in resp_url
                and response.status == 200
            ):
                try:
                    body = await response.json()
                    if body.get("merchant") or body.get("menu"):
                        captured_data["result"]     = body
                        captured_data["source_url"] = resp_url
                except Exception:
                    pass

        page.on("response", handle_response)

        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            deadline = time.time() + 15
            while time.time() < deadline:
                if "result" in captured_data:
                    break
                await asyncio.sleep(0.3)
            if "result" not in captured_data:
                await page.evaluate("window.scrollBy(0, 300)")
                await asyncio.sleep(3)
        except PlaywrightTimeout:
            await browser.close()
            return {"status": "error", "message": f"Timeout setelah {timeout_ms//1000}s"}
        except Exception as e:
            await browser.close()
            return {"status": "error", "message": str(e)}

        await browser.close()

        if "result" in captured_data:
            return {"status": "ok", "data": captured_data["result"]}
        return {"status": "error", "message": "API tidak terpanggil — URL mungkin tidak valid."}


def run_playwright(url: str, lat: float, lng: float, timeout_s: int = 30) -> dict:
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(
            scrape_with_playwright(url, lat, lng, timeout_ms=timeout_s * 1000)
        )
        loop.close()
        return result
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ─────────────────────────────────────────────
# REVERSE GEOCODING (Nominatim / OpenStreetMap)
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def reverse_geocode(lat: float, lng: float) -> str:
    """Konversi koordinat ke alamat teks pakai Nominatim (gratis, no API key)."""
    if not lat or not lng:
        return ""
    try:
        params = urllib.parse.urlencode({"lat": lat, "lon": lng, "format": "json"})
        url    = f"https://nominatim.openstreetmap.org/reverse?{params}"
        req    = urllib.request.Request(url, headers={"User-Agent": "GrabFoodScraper/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = _json.loads(resp.read().decode())
        return data.get("display_name", "")
    except Exception:
        return f"{lat}, {lng}"


# ─────────────────────────────────────────────
# PARSE JSON → INFO RESTO + MENU
# Disesuaikan dengan struktur JSON aktual GrabFood
# ─────────────────────────────────────────────
def parse_data(raw: dict, source_url: str) -> dict:
    """
    Struktur JSON aktual GrabFood:
      raw.merchant.name
      raw.merchant.latlng.latitude / longitude
      raw.merchant.menu.categories[].items[]   ← menu ada di DALAM merchant!

    Fallback: raw.menu.categories (jika menu ada di top-level)
    Contoh: priceInMinorUnit = 3990000 → Rp 39.900 (÷100)
    """
    # ── Unwrap berbagai kemungkinan struktur JSON wrapper ─────────────
    if not raw.get("merchant"):
        for wrapper_key in ("data", "result", "payload", "response"):
            inner = raw.get(wrapper_key)
            if isinstance(inner, dict) and inner.get("merchant"):
                raw = inner
                break

    merchant = raw.get("merchant") or {}

    nama_resto = merchant.get("name", "Unknown Restaurant")
    # Koordinat: coba semua variasi nama field
    latlng = (
        merchant.get("latLng")
        or merchant.get("latlng")
        or merchant.get("lating")
        or {}
    )
    lat = latlng.get("latitude", "")
    lng = latlng.get("longitude", "")

    menu_list = []
    # PENTING: menu ada di DALAM merchant (bukan di top-level raw)
    # Fallback ke raw.menu jika tidak ada di merchant
    menu_obj   = merchant.get("menu") or raw.get("menu") or {}
    categories = menu_obj.get("categories", []) if isinstance(menu_obj, dict) else []

    for cat in categories:
        cat_name = cat.get("name", "")
        for item in cat.get("items", []):
            name = item.get("name", "").strip()
            if not name:
                continue

            deskripsi = (item.get("description") or "").strip()

            # Prioritas ambil harga: priceInMinorUnit → discountedPriceInMin → takeawayPriceInMin
            price_minor = (
                item.get("priceInMinorUnit")
                or item.get("discountedPriceInMin")
                or item.get("discountedTakeawayPriceInMin")
                or item.get("takeawayPriceInMin")
                or 0
            )

            try:
                # priceInMinorUnit = rupiah × 100, jadi ÷100 untuk dapat Rupiah
                price_rupiah = int(price_minor) // 100
            except (TypeError, ValueError):
                price_rupiah = None

            # Format Rupiah: Rp 39.900 (titik sebagai pemisah ribuan)
            harga_str = (
                "Rp {:,}".format(price_rupiah).replace(",", ".")
                if price_rupiah else "N/A"
            )

            menu_list.append({
                "Kategori":    cat_name,
                "Nama Menu":   name,
                "Deskripsi":   deskripsi,
                "Harga (Rp)":  price_rupiah,   # integer, untuk kalkulasi
                "Harga":       harga_str,       # string tampilan
            })

    # Reverse geocoding koordinat → alamat
    alamat = reverse_geocode(lat, lng) if (lat and lng) else ""

    return {
        "url":        source_url,
        "nama_resto": nama_resto,
        "lat":        lat,
        "lng":        lng,
        "alamat":     alamat,
        "menu":       menu_list,
    }


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────
THIN         = Side(style="thin", color="000000")
BLACK_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN_FILL   = PatternFill("solid", fgColor="00b14f")
GRAY_FILL    = PatternFill("solid", fgColor="D9D9D9")
YELLOW_FILL  = PatternFill("solid", fgColor="FFF9C4")


def auto_fit_columns(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border    = BLACK_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_header(ws, header_row=1, fill=GREEN_FILL, font_color="FFFFFF"):
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = fill


def safe_sheet_name(name, existing_names):
    name = re.sub(r'[\\/*?:\[\]]', '', name)[:31] or "Sheet"
    base, counter = name, 1
    while name in existing_names:
        suffix = f"_{counter}"
        name   = base[:31 - len(suffix)] + suffix
        counter += 1
    return name


def fmt_rupiah(val):
    return "Rp {:,}".format(val).replace(",", ".") if val else "N/A"


def build_excel(parsed_list: list) -> bytes:
    """
    Layout Excel per sheet restoran:
      Baris 1 : URL restoran
      Baris 2 : Nama Restoran
      Baris 3 : Koordinat (lat, lng)
      Baris 4 : Header tabel menu  ← warna hijau
      Baris 5+: Data menu (Kategori | Nama Menu | Harga)

    Sheet SUMMARY (paling kiri):
      Nama Restoran | URL | Latitude | Longitude | Total Menu | Rata-rata | Min | Max
    """
    output       = io.BytesIO()
    sheet_names  = []
    summary_rows = []
    MENU_COLS    = ["Kategori", "Nama Menu", "Deskripsi", "Harga"]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for parsed in parsed_list:
            menu_list  = parsed["menu"]
            nama_resto = parsed["nama_resto"]

            df = (
                pd.DataFrame(menu_list, columns=MENU_COLS)
                if menu_list else pd.DataFrame(columns=MENU_COLS)
            )
            sheet_name = safe_sheet_name(nama_resto, sheet_names)
            sheet_names.append(sheet_name)
            # Tabel menu mulai di baris 6 (startrow=5 → row Excel ke-6, karena 4 baris info di atas)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)

            prices = [m["Harga (Rp)"] for m in menu_list if m.get("Harga (Rp)")]
            summary_rows.append({
                "Nama Restoran":   nama_resto,
                "URL":             parsed["url"],
                "Alamat":          parsed.get("alamat", ""),
                "Latitude":        parsed["lat"],
                "Longitude":       parsed["lng"],
                "Total Menu":      len(menu_list),
                "Rata-rata Harga": fmt_rupiah(sum(prices)//len(prices) if prices else None),
                "Harga Terendah":  fmt_rupiah(min(prices) if prices else None),
                "Harga Tertinggi": fmt_rupiah(max(prices) if prices else None),
            })

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="SUMMARY", index=False)

    output.seek(0)
    wb = load_workbook(output)

    for i, parsed in enumerate(parsed_list):
        ws = wb[sheet_names[i]]

        # ── 4 baris info restoran di atas tabel ──────────────────────────
        alamat_str = parsed.get("alamat") or ""
        info_rows = [
            ("URL",           parsed["url"]),
            ("Nama Restoran", parsed["nama_resto"]),
            ("Alamat",        alamat_str),
            ("Koordinat",     f"{parsed['lat']}, {parsed['lng']}"),
        ]
        for r, (label, value) in enumerate(info_rows, start=1):
            ws.cell(row=r, column=1, value=label).font  = Font(bold=True)
            ws.cell(row=r, column=1).fill               = YELLOW_FILL
            ws.cell(row=r, column=1).border             = BLACK_BORDER
            ws.cell(row=r, column=1).alignment          = Alignment(vertical="center")
            ws.cell(row=r, column=2, value=value).fill  = YELLOW_FILL
            ws.cell(row=r, column=2).border             = BLACK_BORDER
            ws.cell(row=r, column=2).alignment          = Alignment(vertical="center", wrap_text=True)

        # Baris kosong ke-5 sebagai spacer (bawaan dari to_excel startrow=5)
        apply_borders(ws)
        style_header(ws, header_row=6, fill=GREEN_FILL)   # header tabel di baris ke-6
        auto_fit_columns(ws)
        ws.freeze_panes = "A7"

    ws_sum = wb["SUMMARY"]
    apply_borders(ws_sum)
    style_header(ws_sum, header_row=1, fill=GRAY_FILL, font_color="000000")
    auto_fit_columns(ws_sum)
    ws_sum.freeze_panes = "A2"
    wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames) + 1)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


# ─────────────────────────────────────────────
# HELPERS UI
# ─────────────────────────────────────────────
def stat_card(label, value, small=False):
    fs = "0.95rem" if small else "1.3rem"
    return (
        f'<div class="stat-card">'
        f'<div class="label">{label}</div>'
        f'<div class="value" style="font-size:{fs}">{value}</div>'
        f'</div>'
    )


# ─────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────
def main():
    if not PLAYWRIGHT_AVAILABLE:
        st.error(
            "❌ **Playwright tidak terinstall.**\n\n"
            "Jalankan di VPS:\n"
            "```\npip install playwright\n"
            "playwright install chromium\n"
            "playwright install-deps chromium\n```"
        )
        return

    st.markdown('<div class="main-title">GrabFood Menu Scraper 🎭</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-title">Versi Playwright — browser sungguhan di background, '
        'tidak perlu login atau input cookie</div>',
        unsafe_allow_html=True,
    )
    st.markdown("""
    <div class="info-box">
        📌 <strong>Cara pakai:</strong><br>
        1. Paste URL restoran GrabFood (satu per baris)<br>
        2. Klik <strong>Mulai Scraping</strong><br>
        <small>⏱️ ~10–20 detik per restoran karena membuka browser sungguhan.</small>
    </div>
    """, unsafe_allow_html=True)

    urls_input = st.text_area(
        "🔗 Link Restoran GrabFood (satu URL per baris)",
        placeholder=(
            "https://food.grab.com/id/en/restaurant/go-steak-joglo-delivery/6-C6MCN2JDYC1AT\n"
            "https://food.grab.com/id/en/restaurant/nama-restoran-delivery/6-XXXXXXXX"
        ),
        height=130,
    )

    with st.expander("📍 Lokasi default (dipakai jika koordinat tidak terdeteksi)"):
        c1, c2 = st.columns(2)
        with c1:
            lat = st.number_input("Latitude",  value=-6.2088,   format="%.6f")
        with c2:
            lng = st.number_input("Longitude", value=106.8456, format="%.6f")

    with st.expander("⚙️ Pengaturan"):
        timeout_s = st.slider("Timeout per restoran (detik)", 15, 60, 30, 5)

    col_btn, col_debug = st.columns([2, 1])
    with col_btn:
        run_btn = st.button("🚀 Mulai Scraping")
    with col_debug:
        show_debug = st.toggle("🐛 Debug JSON", value=False)

    if not run_btn:
        return

    raw_urls = [u.strip() for u in urls_input.strip().splitlines() if u.strip()]
    if not raw_urls:
        st.error("⚠️ Masukkan minimal satu URL.")
        return

    overall_text   = st.empty()
    overall_bar    = st.progress(0)
    detail_text    = st.empty()
    parsed_results = []
    failed_urls    = []
    total          = len(raw_urls)

    for idx, url in enumerate(raw_urls):
        overall_text.info(f"🔄 Scraping restoran **{idx + 1} / {total}**...")
        overall_bar.progress(int(idx / total * 100))
        detail_text.caption(f"  ↳ Membuka browser untuk: `{url[:70]}`")

        pw_result = run_playwright(url, lat, lng, timeout_s)

        if pw_result["status"] == "error":
            st.warning(f"⚠️ **Gagal:** `{pw_result['message']}`\nURL: {url}")
            failed_urls.append(url)
            continue

        raw_data = pw_result["data"]

        if show_debug:
            with st.expander(f"🐛 Raw JSON — {url[:50]}..."):
                # Tampilkan top-level keys untuk diagnosa struktur
                st.caption(f"Top-level keys: `{list(raw_data.keys())}`")
                st.json(raw_data)

        parsed = parse_data(raw_data, url)
        parsed_results.append(parsed)

        menu_list = parsed["menu"]
        emoji = "✅" if menu_list else "⚠️"
        alamat_short = (parsed.get("alamat") or f"{parsed['lat']}, {parsed['lng']}")[:70]
        detail_text.caption(
            f"  {emoji} **{parsed['nama_resto']}** "
            f"| 📍 {alamat_short} "
            f"| {len(menu_list)} item menu"
        )

        if idx < total - 1:
            time.sleep(1.5)

    overall_bar.progress(100)
    overall_text.success(f"✅ Selesai! {len(parsed_results)}/{total} restoran berhasil.")

    if failed_urls:
        st.warning("⚠️ URL berikut gagal:\n" + "\n".join(f"- {u}" for u in failed_urls))

    if not parsed_results:
        st.error("❌ Tidak ada data yang berhasil di-scrape.")
        return

    # ── Tampilkan Hasil ───────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Hasil Scraping")
    total_items_all = sum(len(p["menu"]) for p in parsed_results)

    if len(parsed_results) > 1:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(stat_card("Total Restoran", len(parsed_results)), unsafe_allow_html=True)
        with c2:
            st.markdown(stat_card("Total Item Menu", total_items_all), unsafe_allow_html=True)

    for parsed in parsed_results:
        menu_list  = parsed["menu"]
        nama_resto = parsed["nama_resto"]

        st.markdown(f'<div class="resto-header">🏪 {nama_resto}</div>', unsafe_allow_html=True)
        alamat_ui = parsed.get("alamat") or f"{parsed['lat']}, {parsed['lng']}"
        st.caption(f"📍 {alamat_ui}  |  🔗 {parsed['url'][:70]}")

        if not menu_list:
            st.warning(f"⚠️ Menu kosong untuk **{nama_resto}**.")
            continue

        prices = [m["Harga (Rp)"] for m in menu_list if m.get("Harga (Rp)")]
        avg  = fmt_rupiah(sum(prices)//len(prices) if prices else None)
        low  = fmt_rupiah(min(prices) if prices else None)
        high = fmt_rupiah(max(prices) if prices else None)

        c1, c2, c3, c4 = st.columns(4)
        with c1: st.markdown(stat_card("Total Menu",  len(menu_list)),   unsafe_allow_html=True)
        with c2: st.markdown(stat_card("Rata-rata",   avg,  small=True), unsafe_allow_html=True)
        with c3: st.markdown(stat_card("Terendah",    low,  small=True), unsafe_allow_html=True)
        with c4: st.markdown(stat_card("Tertinggi",   high, small=True), unsafe_allow_html=True)

        with st.expander(f"📋 Preview menu ({len(menu_list)} item)"):
            df = pd.DataFrame(menu_list, columns=["Kategori", "Nama Menu", "Deskripsi", "Harga"])
            st.dataframe(df, use_container_width=True, height=300)

    # ── Download Excel ─────────────────────────────────────────────────────
    st.markdown("---")
    valid = [p for p in parsed_results if p["menu"]]
    if valid:
        with st.spinner("📊 Membuat file Excel..."):
            excel_bytes = build_excel(valid)
        n        = len(valid)
        filename = "grabfood_menu.xlsx" if n == 1 else f"grabfood_menu_{n}_resto.xlsx"
        st.download_button(
            label=f"⬇️ Download Excel ({n} restoran, {total_items_all} item)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.error("❌ Tidak ada data untuk di-export.")


if __name__ == "__main__":
    main()