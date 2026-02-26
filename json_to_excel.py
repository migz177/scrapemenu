import json
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Config ───────────────────────────────────────────────────────────────────
JSON_DIR    = os.path.join(os.path.dirname(__file__), "10resto")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "10resto_menus.xlsx")
# ─────────────────────────────────────────────────────────────────────────────

THIN_BLACK = Side(style="thin", color="000000")
BLACK_BORDER = Border(
    left=THIN_BLACK, right=THIN_BLACK,
    top=THIN_BLACK,  bottom=THIN_BLACK,
)


def load_json(filepath: str) -> dict:
    with open(filepath, encoding="utf-8") as f:
        return json.load(f)


def auto_fit_columns(ws):
    """Set column width based on the longest content in each column."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Cap width at 80 to avoid extremely wide columns (e.g. long descriptions)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


def apply_borders(ws):
    """Apply thin black border to all used cells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = BLACK_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_header_row(ws, header_row: int = 1):
    """Bold the header row."""
    for cell in ws[header_row]:
        cell.font = Font(bold=True)


def main():
    json_files = sorted(glob.glob(os.path.join(JSON_DIR, "*.json")))
    if not json_files:
        print(f"Tidak ada file JSON ditemukan di folder: {JSON_DIR}")
        return

    print(f"Ditemukan {len(json_files)} file JSON:")
    for f in json_files:
        print(f"  • {os.path.basename(f)}")

    summary_rows = []

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        # ── Sheet per restoran ────────────────────────────────────────────────
        for filepath in json_files:
            data       = load_json(filepath)
            resto_name = os.path.splitext(os.path.basename(filepath))[0].capitalize()
            url        = data.get("url", "")

            menu_rows = [
                {
                    "Nama Menu" : item.get("name", ""),
                    "Deskripsi" : item.get("description", ""),
                    "Harga"     : item.get("price_formatted", ""),
                }
                for item in data.get("menus", [])
            ]

            df = pd.DataFrame(menu_rows)
            sheet_name = resto_name[:31]

            # Tulis dengan startrow=2 agar ada ruang untuk info URL di baris 1
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

            summary_rows.append({"Restoran": resto_name, "URL": url})

        # ── Sheet SUMMARY ─────────────────────────────────────────────────────
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="SUMMARY", index=False)

    # ── Post-process: tambah URL info + styling ───────────────────────────────
    wb = load_workbook(OUTPUT_FILE)

    for filepath in json_files:
        data       = load_json(filepath)
        resto_name = os.path.splitext(os.path.basename(filepath))[0].capitalize()
        url        = data.get("url", "")
        ws         = wb[resto_name[:31]]

        # Tulis info URL di baris 1
        ws["A1"] = "URL"
        ws["B1"] = url
        ws["A1"].font = Font(bold=True)

        # Baris 2 = kosong (sudah kosong karena startrow=2), baris 3 = header kolom
        apply_borders(ws)
        style_header_row(ws, header_row=3)  # header dari pandas ada di baris 3
        auto_fit_columns(ws)
        ws.freeze_panes = "A4"  # freeze di bawah header

    # ── Styling SUMMARY ───────────────────────────────────────────────────────
    ws_sum = wb["SUMMARY"]
    apply_borders(ws_sum)
    style_header_row(ws_sum, header_row=1)
    auto_fit_columns(ws_sum)
    ws_sum.freeze_panes = "A2"

    # Pindahkan SUMMARY ke posisi pertama
    wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames) + 1)
    wb.save(OUTPUT_FILE)

    print(f"\n✅ File Excel berhasil dibuat: {OUTPUT_FILE}")
    print(f"   Sheet: SUMMARY + {len(json_files)} sheet restoran")


if __name__ == "__main__":
    main()
